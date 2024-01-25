
from selenium import webdriver
import openpyxl
import time

# Initialize the Chrome WebDriver (provide the path to your WebDriver executable)
driver = webdriver.Chrome()

def get_image_url_from_google(search_query):
    search_url = f"https://www.google.com/search?hl=en&tbm=isch&q={search_query}"
    driver.get(search_url)

    # Wait for images to load
    time.sleep(2)  # sleep for a few seconds to allow the page to load

    # Find the images after the JavaScript has loaded them
    images = driver.find_elements_by_css_selector('img.rg_i')

    # Return the first image URL
    if images:
        return images[0].get_attribute('src')

# Load the Excel workbook
workbook = openpyxl.load_workbook('test-file.xlsx')
sheet = workbook.active

# Fetch image URLs and add them to the Excel file
for row in range(2, sheet.max_row + 1):
    search_term = sheet.cell(row, 1).value
    if search_term:
        image_url = get_image_url_from_google(search_term)
        sheet.cell(row, 2).value = image_url

# Close the WebDriver
driver.quit()

# Save the Excel workbook
workbook.save('my_updated_file.xlsx')


